import openpyxl
import smtplib
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText

workbook=openpyxl.load_workbook("attendance.xlsx")
sheet=workbook.active

def savefile():
    workbook.save('attendance.xlsx')
    print("Saved Sucessfully")


def create_details():
    sheet['A1']="SNO"
    sheet['B1']="Student_Name"
    sheet['C1']="Roll_NO"
    sheet['D1']="Subject_Name"
    sheet['E1']="Department"
    sheet['F1']="Attendance"

    sub_name = input("Enter the name of the subject:")
    roll_no=int(input("Enter the roll of the student:"))
    student_name=input("Enter the name of the Student:") 
    dept=input("Enter the Department:")
    
    sub_column = 4  
    student_column=2
    roll_column=3
   
    next_row=sheet.max_row+1
    sheet.cell(row= next_row, column=sub_column).value = sub_name
    num=1
    sheet.cell(row= next_row,column=1).value=next_row-1

    sheet.cell(row= next_row,column=student_column).value=student_name
    sheet.cell(row= next_row,column=roll_column).value=roll_no
    sheet.cell(row= next_row,column=5).value=dept
    sheet.cell(row= next_row,column=6).value="Absent"

    sub_column += sub_column
    student_column+=student_column
    roll_column+=roll_column

    print("Displaying Details....")
    print("Student Name:",student_name)
    print("Roll No:",roll_no)
    print("Attendance:"+"Absent")
    print("Department:",dept)
    savefile()
    message="You were absent for today "+sub_name+"Class"
    attendance(subject_name=sub_name)
    r=int(input("Want to continue again: --1 for --0:"))
    if(r==1):
     create_details()
    else:
       print("Exiting...")

def SendMail(subject, message):
   try:
      sender_mail="Your Outlook Id"
      password='Your Password'
      reciever_mail=input("Enter the mail id of the student:")  
      Message=message
      m=MIMEText(Message)
      m['subject']=subject
      server=smtplib.SMTP('smtp-mail.outlook.com',587)
      server.starttls()
      server.login(user=sender_mail,password=password)
      print("Login successful")
      server.sendmail(sender_mail,reciever_mail,msg=m.as_string())
      server.quit()
      print("Sent mail successfully")
   except Exception as e:
      print("Error while sending mail")

def attendance(subject_name):
   try:
      threshold=1
      openpyxl.load_workbook("attendance.xlsx")
      sheet=workbook.active
      attendance=sheet.max_row-1
      if(attendance>threshold):
         SendMail("attendance Report","Attendance shortage kindly attend the class regularly...")
      elif(attendance<=threshold):
         SendMail("attendance Report","Warning you are absent for"+subject_name +" Class and u have only one leave left")


   except Exception as e:
      print("error")
             


print("NITPY STUDENTS ATTENDANCE TRACKER SYSTEM")
while True:
 print("1.Create Details 2.SaveFile 3.Sendmail 4.Attendance 5.Exit")
 option=int(input("Enter your option:"))
 if(option==1):
   create_details()

 elif(option==2):
    savefile()


 elif(option==3):
   sub=input("Enter the subject for the mail:")
   body=input("Enter the body of the mail:")
   SendMail(sub,body)


 elif(option==4):
  subname=input("Enter the subject name:")
  attendance(subject_name=subname)
  
 elif(option==5):
    print("Exiting....")
    break

   
   
 







   
   


    


    

